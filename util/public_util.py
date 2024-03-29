import os
import csv
import json
import time
import xlrd
import xlwt
import base64
import random
import pymysql
import openpyxl
import cx_Oracle
from Crypto.Cipher import PKCS1_v1_5 as Cipher_pkcs1_v1_5
from Crypto.Signature import PKCS1_v1_5 as Signature_pkcs1_v1_5
from Crypto.PublicKey import RSA
from Crypto.Hash import MD5
from Crypto.Hash import SHA


def random_text_base_date(pre: str = None, suffix: str = None):
    """
    根据输入的前缀和后缀，随机生成内容
    :param pre: 前缀
    :param suffix: 后缀
    :return:
    """
    date_text = time.strftime('%Y_%m_%d_%H_%M')
    if pre is None and suffix is None:
        return date_text
    elif pre is not None:
        if suffix is not None:
            return pre + '_' + date_text + '_' + suffix
        else:
            return pre + '_' + date_text
    else:
        if suffix is not None:
            return date_text + '_' + suffix


def get_product_info_on_performance(store_no: list, file: str, env: str):
    """
    通过门店no查询当前门店下的产品信息， 包括：商品id、规格id、快照版本id、属性id、属性选项id
    :param store_no:
    :param file:
    :return:
    """

    if env == 'uat':
        # UAT环境
        select_mysql_env = MySQL(user='lifekh_takeaway_uat', password='lifekh_takeaway_uat_2020',
                             host='172.16.27.10', port=3400, database='lifekh_takeaway_uat')
        database = 'lifekh_takeaway_uat'
    elif env == 'pro':
        # 金边机房
        select_mysql = MySQL(user='lifekh_takeaway_query', password='q1h9MqKpgX5V9qVFfFjEC',
                             host='10.24.255.42', port=3300, database='lifekh_takeaway')
        database = 'lifekh_takeaway'
    elif env == '310':
        # 金边310机房
        select_mysql = MySQL(user='lifekh_takeaway_query4hei', password='h9X5V9qMqKpgX5fFjECMgqFVFj',
                             host='10.34.255.26', port=3300, database='lifekh_takeaway')
        database = 'lifekh_takeaway'

    result = []
    for i in range(len(store_no)):
        # 根据门店号查询门店下的所有商品id，输出对应的set集合
        select_product_id_list = select_mysql.select(
            "SELECT id as 商品id FROM `{database}`.`product` WHERE `store_no` = '{store_no}' and `del_state` = 10"
            .format(database=database, store_no=store_no[i]))

        product_id_list = []
        for n in range(len(select_product_id_list)):
            # 将查询 - 商品id 单独添加的对应的list中
            product_id_list.append(select_product_id_list[n]['商品id'])
        # print("商品id："+str(product_id_list))

        for m in range(len(product_id_list)):
            # 根据商品id查询商品SKU id，输出对应的set集合
            select_product_specification_id_list = select_mysql.select(
                "SELECT id as 规格id FROM `{database}`.`product_specification` WHERE `product_id` = '{product_id}'"
                .format(database=database, product_id=product_id_list[m]))
            product_specification_id_list = []  # 商品SKU id list

            for k in range(len(select_product_specification_id_list)):
                # 将查询 - 规格id 单独添加的对应的list中
                product_specification_id_list.append((select_product_specification_id_list[k]['规格id']))

                select_product_multiple_version_id_list = select_mysql.select(
                    "SELECT id as 快照版本id FROM `{database}`.`product_multiple_version` WHERE `product_id` = '{product_id}' order by `update_time` Desc limit 0,1"
                    .format(database=database, product_id=product_id_list[m]))
                product_multiple_version_id_list = []  # 商品快照版本id list
                for q in range(len(select_product_multiple_version_id_list)):
                    # 将查询 - 快照版本id 单独添加的对应的list中
                    product_multiple_version_id_list.append(select_product_multiple_version_id_list[q]['快照版本id'])

                    select_product_property_id_list = select_mysql.select(
                        "SELECT id, required_selection FROM `{database}`.`product_property` WHERE `product_id` = '{product_id}'"
                            .format(database=database, product_id=product_id_list[m]))
                    product_property_id_list = []  # 商品属性id list
                    if len(select_product_property_id_list) > 0:
                        for x in range(len(select_product_property_id_list)):
                            # 将查询 - 属性id 单独添加的对应的list中
                            print(select_product_property_id_list[x]['required_selection'])
                            if select_product_property_id_list[x]['required_selection'] > 1:
                                print('必选属性选项id数>1，不作为性能批量数据中')
                            else:
                                print('必选属性选项id数<=1')
                                product_property_id_list.append(select_product_property_id_list[x]['id'])

                            # 此处判断逻辑，剔除属性id，必选项参数>1的参数
                            if x < len(product_property_id_list):
                                select_product_property_selection_id_list = select_mysql.select(
                                    "SELECT id as 属性选项id FROM `{database}`.`product_property_selection` WHERE `product_property_id` = '{product_property_id}'"
                                        .format(database=database, product_property_id=product_property_id_list[x]))
                                product_property_selection_id_list = []  # 商品属性选项id list
                                for l in range(len(select_product_property_selection_id_list)):
                                    # 将查询 - 商品属性选项id 单独添加的对应的list中
                                    product_property_selection_id_list.append(select_product_property_selection_id_list[l]['属性选项id'])
                                    print('商品id：'+str(product_id_list[m]) +
                                          ',规格id：'+str(product_specification_id_list[k]) +
                                          ',快照版本id：'+str(product_multiple_version_id_list[q]) +
                                          ',属性id：'+str(product_property_id_list[x]) +
                                          ',属性选项id：'+str(product_property_selection_id_list[l]))
                                    result.append(str(store_no[i]) +
                                                  ','+str(product_id_list[m]) +
                                                  ','+str(product_specification_id_list[k]) +
                                                  ','+str(product_multiple_version_id_list[q]) +
                                                  ','+str(product_property_id_list[x]) +
                                                  ','+str(product_property_selection_id_list[l]))
                            else:
                                print('当前商品id：{},有属性id：{}，但属性必选项参数>1，不添加到性能数据中去'.format(str(product_id_list[m]), str(select_product_property_id_list[x]['id'])))

                    else:
                        print('当前商品id：{},无属性id'.format(str(product_id_list[m])))
                        print('商品id：' + str(product_id_list[m]) +
                              ',规格id：' + str(product_specification_id_list[k]) +
                              ',快照版本id：' + str(product_multiple_version_id_list[q]))
                        result.append(str(store_no[i]) +
                                      ',' + str(product_id_list[m]) +
                                      ',' + str(product_specification_id_list[k]) +
                                      ',' + str(product_multiple_version_id_list[q])
                                      )

        first_line_data = 'storeNo,goodsId,goodsSkuId,inEffectVersionId,propertyId,properties'
        write_csv_product_info(file=file, data=result, first_line_data=first_line_data)


def write_csv_product_info(file: str, data: list, first_line_data: str):
    """

    :param file: 写入文件目录，需包含文件本身
    :param data: 需要写入的数据
    :param first_line_data:
    :return:
    """
    with open(file, newline='', encoding='utf-8', mode='w') as f:
        csv_writer = csv.writer(f)
        csv_writer.writerow([first_line_data])
        for i in range(len(data)):
            data_info = data[i]
            csv_writer.writerow([data_info])
        print('此方法输出csv文件需要手动替换""，否则jmeter中执行会报错！！！！！')


def read_txt(file: str, key_word: str):
    """
    读取接口报错中相关关键值的键值
    :param file:
    :param key_word:
    :return:
    """
    with open(file, encoding='utf-8', mode='r') as f:
        result = []
        for i in f.readlines():
            if key_word in i:
                i = i.strip().replace('"', '').replace('\\n', '').replace(key_word + ':', '').replace(',', '').strip()
                result.append(i)
                print(i)
    return result


def get_phone_number_cambodia(prefix: bool = True, check: bool = False, env: str = 'uat'):
    """
    随机生成855可用运营商号段号码
    :param prefix:是否需要855前置，默认值：True
    :param check:是否需要检验随机生成的号码注册使用过
    :return:
    """
    phone_list = ['11', '12', '14', '17', '61', '76', '77', '78', '79', '85', '89', '92',
                  '95', '99', '10', '15', '16', '69', '70', '81', '86', '87', '93',
                  '96', '98', '31', '60', '66', '67', '68', '71', '88', '90', '97',
                  '13', '80', '83', '84', '38', '18'
                  ]  # 柬埔寨可正常使用手机号号段
    phone_seven_length = ['76', '96', '31', '71', '88', '97', '38', '18']
    phone_segment = random.choice(phone_list)
    if phone_segment == '12':
        num_length = random.randint(6, 7)
    elif phone_segment in phone_seven_length:
        num_length = random.randint(7, 7)
    else:
        num_length = random.randint(6, 6)
    if prefix:
        register_number = '8550' + phone_segment + "".join(random.choice("0123456789") for i in range(num_length))
    else:
        register_number = phone_segment + "".join(random.choice("0123456789") for i in range(num_length))

    # 判断是否需要校验号码已注册，校验会严重降低速度
    if check:
        if env == 'uat':
            database = 'LIFEKH_MP_CUSTOMER_UAT'
            check_oracle = Oracle(username='lifekh_mp_customer_uat', password='lifekh_mp_customer_uat_2020',
                                  address='172.16.27.10:1521/lifekh')
        else:
            database = 'LIFEKH_MP_CUSTOMER'
            check_oracle = Oracle(username='lifekh_mp_customer_uat', password='lifekh_mp_customer_uat_2020',
                                  address='172.16.27.10:1521/lifekh')
        check_data = check_oracle.select("SELECT LOGIN_NAME from `{database}`.`USER_OPERATOR_LOGIN_INFO` WHERE `LOGIN_NAME` ='{loginName}'".format(database=database, loginName=register_number))

        if len(check_data) != 0:
            if register_number in ''.join(check_data[0]):
                print('%s 账号已注册' % register_number)
                get_phone_number_cambodia()
        else:
            print('%s 可正常使用' % register_number)

    return register_number


def get_email(postfix: str = ''):
    """
    生成当前常用的email
    :param postfix:
    :return:
    """
    email_postfix_list = ['@qq.com', '@163.com', '@gmail.com', '@yahoo.com', '@msn.com', '@hotmail.com',
                          '@aol.com', '@live.com', '@0355.net', '@163.net', '@263.net', '@3721.net']
    email_postfix = random.choice(email_postfix_list)
    email_length = random.randint(7, 10)
    email = "".join(random.choice("0123456789") for i in range(email_length)) + email_postfix
    return email


def change_html(source_file_path: str, target_file_path: str):
    """
    目前因为输出的html中，会彩色日志的形式，导致log中存在shell在控制台显示的代码
    此方法用于去除生成多余的代码，并生成新的文件，同时删除旧文件

    :param source_file_path:
    :param target_file_path:
    :return:
    """
    with open(source_file_path, encoding='utf-8', mode='r') as rf:
        with open(target_file_path, mode='a+') as wf:

            for i in rf.readlines():
                if '[' in i:
                    i = i.strip()\
                        .replace('[32m', '')\
                        .replace('[0m', '')\
                        .replace('[33m', '')\
                        .replace('[31m', '')\
                        .replace('[91m', '')
                    wf.write(i)
                    wf.write('<br>')
                else:
                    wf.write(i)

    os.remove(source_file_path)  # 删除文件


def write_csv_loginname(file: str, first_line_data: str = 'loginName', times=1):
    """
    :param file:需要写入的文件路径
    :param first_line_data:写入文件首行需要填写的内容，默认值：loginName
    :param times:循环次数，默认值：1次
    :return:
    """
    with open(file, newline='', encoding='utf-8', mode='w') as f:
        csv_writer = csv.writer(f)
        csv_writer.writerow([first_line_data])
        for i in range(0, times):
            login_no = get_phone_number_cambodia()
            csv_writer.writerow([login_no])


def api_data_dict_exchange_str(data: dict):
    """
    将dict的数据类型，转换为json格式，并且去除空格
    :param data:
    :return:
    """
    data_str = json.dumps(data).replace(' ', '')
    return data_str


def escape_double_quotation_marks(data: str):
    escape_data = ''
    for i in data:
        if i == '"':
            escape_data = escape_data + '\\' + i
        else:
            escape_data = escape_data + i

    return escape_data


def api_query_data(api_url: str, api_data: dict):
    api_data_str = api_data_dict_exchange_str(api_data)
    request_data = '{"apiUrl":"' + api_url + '","apiData":' + api_data_str
    return escape_double_quotation_marks(request_data)


class Operator_xls():
    """

    """

    def create_xls(self, sheet_name: str, file_name: str):
        """

        :param sheet_name:
        :param file_name:
        :return:
        """
        wb = openpyxl.Workbook()
        wb.create_sheet(sheet_name)
        wb.save(file_name)

    def open_xls_by_row(self, file_name: str, row_number: int = 1):
        """
        此方法仅适用于获取xls中加班申请的数据
        :param file_name:
        :param row_number:
        :return:
        """
        wb = xlrd.open_workbook_xls(filename=file_name)
        sh = wb.sheets()[0]
        test_list = []
        for i in range(sh.nrows):
            if i == 0:
                pass
            else:
                temp = sh.row_values(i)[4:-2]
                temp.pop(-4)  # 去除：休假事由 字段
                temp.pop(1)  # 去除：申请日期 字段
                temp.pop(2)  # 去除：职务 字段
                test_list.append(temp)
        # print(test_list)
        return test_list

    def create_write_xls(self, data: list):
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Sheet1')
        for i in range(len(data)):
            for n in range(len(data[i])):
                # print(data[i][n])
                ws.write(i, n, data[i][n])

        wb.save('test_001.xls')


class Rsa_utils():
    """
    RSA非对称加密
    """

    def rsa_sign_by_private_key(self, encryptData, private_key):
        """
        通过RSA私钥加签
        :param encryptData:
        :param private_key:
        :return:
        """
        # privateKey = '-----BEGIN PRIVATE KEY-----\n' + private_key + '\n-----END PRIVATE KEY-----'
        # print(privateKey)
        private_keyBytes = base64.b64decode(private_key)
        priKey = RSA.importKey(private_keyBytes)
        signer = Signature_pkcs1_v1_5.new(priKey)
        data = SHA.new(encryptData.encode('utf-8'))
        signature = signer.sign(data)
        signature = base64.b64encode(signature)
        return signature.decode()

    def rsa_verify_by_public_key(self, encryptData, public_key):
        """
        通过RSA公钥验签
        :param encryptData:
        :param public_key:
        :return:
        """
        # 若输入的key 带有对应的前缀和后缀则不用执行 public_keyBytes = base64.b64decode(public_key)
        # privateKey = '-----BEGIN PRIVATE KEY-----\n' + private_key + '\n-----END PRIVATE KEY-----'
        # print(privateKey)
        public_keyBytes = base64.b64decode(public_key)
        priKey = RSA.importKey(public_keyBytes)
        signer = Signature_pkcs1_v1_5.new(priKey)
        data = MD5.new(encryptData.encode('utf-8'))
        signature = signer.sign(data)
        signature = base64.b64encode(signature)
        return signature.decode()

    def encrypt_by_public_key(self, plaintext, public_key):
        """
        通过RSA公钥加密
        :param plaintext:  明文
        :param public_key: RSA公钥

        :return:           密文
        """
        publicKey = '-----BEGIN PUBLIC KEY-----\n' + public_key + '\n-----END PUBLIC KEY-----'
        print(publicKey)

        rsakey = RSA.importKey(publicKey)
        cipher = Cipher_pkcs1_v1_5.new(rsakey)
        ciphertext = base64.b64encode(cipher.encrypt(plaintext.encode(encoding='UTF-8')))
        return ciphertext.decode('utf8')

    def decrypt_by_private_key(self, ciphertext, private_key):
        """
        通过RSA私钥解密
        :param ciphertext:  密文
        :param private_key: RSA私钥

        :return:            明文
        """
        privateKey = '-----BEGIN PRIVATE KEY-----\n' + private_key + '\n-----END PRIVATE KEY-----'
        rsakey = RSA.importKey(privateKey)
        cipher = Cipher_pkcs1_v1_5.new(rsakey)
        plaintext = cipher.decrypt(base64.b64decode(ciphertext), None)
        return plaintext.decode('utf8')


class Oracle:
    # 设置访问本地oracle_client,若没有可能会报错：
    # DPI-1047: Cannot locate a 64-bit Oracle Client library: "dlopen(libclntsh.dylib, 1): image not found
    cx_Oracle.init_oracle_client(lib_dir=r"/usr/local/lib/instantclient_19_8")

    def __init__(self, username: str = 'lifekh_mp_customer_uat', password: str = 'lifekh_mp_customer_uat_2020',
                 address: str = '172.16.27.10:1521\lifekh'):
        """
        默认连接表：lifekh_mp_customer_uat
        :param username:
        :param password:
        :param address:
        """
        self.username = username
        self.password = password
        self.address = address

    def _connect(self):
        db = cx_Oracle.connect(self.username, self.password, self.address)
        return db

    def _execute(self, sql: str, mode: str):
        db = self._connect()
        cur = db.cursor()
        cur.execute(sql)

        if mode == 'r':
            rows = cur.fetchall()
        else:
            db.commit()
            print('执行sql语句：{}'.format(sql))

        cur.close()
        db.close()
        if mode == 'r':
            return rows

    def select(self, sql: str, mode: str = 'r'):
        if 'select' not in sql.lower():
            raise Exception('执行的SQL语句，不是select相关的语句。')
        else:
            return self._execute(sql=sql, mode=mode)

    def update(self, sql: str, mode='w'):
        if 'update' not in sql.lower():
            raise Exception('执行的SQL语句，不是update相关的语句。')
        else:
            return self._execute(sql=sql, mode=mode)

    def insert(self, sql: str, mode='w'):
        if 'insert' not in sql.lower():
            raise Exception('执行的SQL语句，不是insert相关的语句。')
        else:
            return self._execute(sql=sql, mode=mode)

    def delete(self, sql: str, mode='w'):
        if 'delete' not in sql.lower():
            raise Exception('执行的SQL语句，不是delete相关的语句。')
        else:
            return self._execute(sql=sql, mode=mode)


class MySQL:
    def __init__(self, user: str = 'lifekh_takeaway_uat', password: str = 'lifekh_takeaway_uat_2020',
                 host: str = '172.16.27.10', port: int = 3400, database: str = 'lifekh_takeaway_uat'):
        """

        :param user:
        :param password:
        :param host:
        :param port:
        :param database:
        """
        self.user = user
        self.password = password
        self.host = host
        self.port = port
        self.database = database

    def _connect(self):
        # 创建connect连接
        py = pymysql.connect(user=self.user, password=self.password, host=self.host, port=self.port,
                             database=self.database, charset="utf8", cursorclass=pymysql.cursors.DictCursor)
        # 获取cursor对象
        cs = py.cursor()
        return py, cs

    def _execute(self, sql: str, mode='r'):
        py, cs = self._connect()
        try:
            cs.execute(sql)
            if mode == 'r':
                data = cs.fetchall()
            elif mode == 'w':
                py.commit()
                data = cs.rowcount
        except:
            data = False
            py.rollback()
        py.close()
        return data

    def select(self, sql: str, mode='r'):
        if 'select' not in sql.lower():
            raise Exception('执行的SQL语句，不是select相关的语句。')
        else:
            return self._execute(sql=sql, mode=mode)

    def update(self, sql: str, mode='w'):
        if 'update' not in sql.lower():
            raise Exception('执行的SQL语句，不是update相关的语句。')
        else:
            return self._execute(sql=sql, mode=mode)

    def insert(self, sql: str, mode='w'):
        if 'insert' not in sql.lower():
            raise Exception('执行的SQL语句，不是insert相关的语句。')
        else:
            return self._execute(sql=sql, mode=mode)

    def delete(self, sql: str, mode='w'):
        if 'delete' not in sql.lower():
            raise Exception('执行的SQL语句，不是delete相关的语句。')
        else:
            return self._execute(sql=sql, mode=mode)


if __name__ == '__main__':
    # test_list = ['99.0.4844.51', '101.0.4951.41']
    # print(get_list_last_driver_version(test_list))
    print(get_email())
    # get_product_info_on_performance(['MS1320194834269442048'], '/Users/windy/Desktop/jmeter_script/chaoA_performance_test/uat_data_info/uat_store_info.csv', 'uat')
    # get_product_info_on_performance(['MS1461534823757172736'], '/Users/windy/Desktop/jmeter_script/chaoA_performance_test/pro310_data_info/pro310_store_info.csv', '310')
    # api_query_data(api_url='https://boss-uat.lifekh.com/boss_web/config/banner/v2/deleteCard.do', api_data=test_dict)

    # print(get_phone_number_cambodia(check=True))
    # change_html('/Users/windy/Documents/code/myself/ui/pytest_ui_web_autotest/test_report/2021_10_15_17_45_39/report.html',
    #             '/Users/windy/Documents/code/myself/ui/pytest_ui_web_autotest/test_report/2021_10_15_17_45_39/report_new.html')
    # test_data = [{'order_no': '1446671320510849024'}]
    # print(test_data[0]['order_no'])
    # test = api_data_dict_exchange_str({"12": 12})
    # mysql = MySQL()
    # sql = "SELECT id, required_selection FROM `lifekh_takeaway_uat`.`product_property` WHERE `product_id` = '{}'".format('1262170')
    # sql_select_data = mysql.select(sql)
    # for i in range(len(sql_select_data)):
    #     print(sql_select_data[i])
    #     if sql_select_data[i]['required_selection'] > 1:
    #         print('必选属性选项id数>1')
    #     else:
    #         print('必选属性选项id数<=1')
    # print(sql_select_data)
    # #
    # test_loginName = write_csv_loginname(file='/Users/windy/Desktop/jmeter_script/chaoA_performance_test/uat_data_info/uat_new_user.csv',
    #                                      times=100)
    # test_rsa = Rsa_utils()
    # temp_1 = test_rsa.encrypt_by_public_key("test", "MIGfMA0GCSqGSIb3DQEBAQUAA4GNADCBiQKBgQCjblMlMerHC2Dx2+mbrW2jjmBDgDaSolsXZuC3e2oc94xBHPn0C4Dyna1xRkBPPWW1I4GoBWRL32LDQ8xqfugv8UxcAiOQdTXVJ5XbHSXCBfA+aySRg9tjfVONyxp1cwUhon23NkdY9Jy/XE7VKUzjCvhwasfbjg83oQY8dyQefQIDAQAB")
    # print(temp_1)
    #
    # a = read_txt('/Users/windy/Desktop/error.txt', 'loginName')
    # lala = input('请输入需要获取的xls文件路径')
    # print(lala)
    # xls = Operator_xls()
    # test = xls.open_xls_by_row(file_name='/Users/windy/Desktop/S-app(23).xls', row_number=1)
    # need_list = rest_by_month('2020', '12', test)
    # xls.create_write_xls(need_list)








